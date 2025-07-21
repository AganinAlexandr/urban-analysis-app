"""
Модуль для обработки Excel файлов с отзывами городской среды
"""

import pandas as pd
import os
import logging
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class ExcelProcessor:
    """Класс для обработки Excel файлов с отзывами"""
    
    def __init__(self):
        """Инициализация процессора Excel"""
        # Определяем обязательные поля для обработки
        self.required_fields_processing = ['group', 'review_text']
        
        # Определяем обязательные поля для сохранения в архив
        self.required_fields_archive = ['group', 'name', 'address', 'review_text', 'date']
        
        # Определяем опциональные поля
        self.optional_fields = [
            'determined_group', 'user_name', 'rating', 'answer_text', 'latitude', 'longitude', 'district',
            'rating_object', 'review_count_from_api', 'review_count_fetched'
        ]
        
        # Маппинг полей для совместимости с разными форматами Excel
        self.field_mapping = {
            # Поля из разных источников
            'object_name': 'name',
            'review_rating': 'rating',
            'review_date': 'date',
            'object_rating': 'rating_object',
            'review_count': 'review_count_from_api',
            'star': 'rating',
            'stars': 'rating',
            'average_rating': 'rating_object',
            'source': 'source'
        }
        
        # Все поддерживаемые поля
        self.supported_fields = self.required_fields_archive + self.optional_fields + ['source']
        
        # Стандартный порядок полей в архиве
        self.archive_field_order = [
            'group', 'determined_group', 'name', 'address', 'review_text', 'date', 'user_name',
            'rating', 'answer_text', 'latitude', 'longitude', 'district',
            'rating_object', 'review_count_from_api', 'review_count_fetched',
            'source', 'sentiment', 'sentiment_score', 'review_type', 'positive_words_count',
            'negative_words_count', 'hash_key'
        ]
    
    def get_excel_info(self, file_path: str) -> Dict:
        """
        Получение информации об Excel файле
        
        Args:
            file_path: Путь к Excel файлу
            
        Returns:
            Словарь с информацией о файле
        """
        try:
            # Загружаем книгу для получения информации о листах
            workbook = load_workbook(file_path, read_only=True)
            
            # Получаем список листов
            sheet_names = workbook.sheetnames
            
            # Получаем информацию о каждом листе
            sheets_info = {}
            for sheet_name in sheet_names:
                try:
                    # Читаем только заголовки для определения колонок
                    df_sample = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)
                    
                    # Применяем маппинг полей
                    mapped_columns = []
                    for col in df_sample.columns:
                        mapped_columns.append(self.field_mapping.get(col, col))
                    
                    # Определяем поддерживаемые поля после маппинга
                    supported_fields = [col for col in mapped_columns if col in self.supported_fields]
                    
                    sheets_info[sheet_name] = {
                        'columns': df_sample.columns.tolist(),
                        'mapped_columns': mapped_columns,
                        'sample_rows': len(df_sample),
                        'supported_fields': supported_fields
                    }
                except Exception as e:
                    logger.warning(f"Ошибка чтения листа {sheet_name}: {str(e)}")
                    sheets_info[sheet_name] = {
                        'columns': [],
                        'sample_rows': 0,
                        'supported_fields': []
                    }
            
            workbook.close()
            
            return {
                'file_path': file_path,
                'file_size': os.path.getsize(file_path),
                'sheets': sheets_info,
                'total_sheets': len(sheet_names)
            }
            
        except Exception as e:
            logger.error(f"Ошибка получения информации об Excel файле {file_path}: {str(e)}")
            return {
                'file_path': file_path,
                'error': str(e),
                'sheets': {},
                'total_sheets': 0
            }
    
    def clean_text_field(self, text: str) -> str:
        """
        Очистка текстового поля от проблемных символов
        
        Args:
            text: Исходный текст
            
        Returns:
            Очищенный текст
        """
        if pd.isna(text) or text is None:
            return ""
        
        text = str(text)
        
        # Удаляем лишние пробелы в начале и конце
        text = text.strip()
        
        # Заменяем множественные пробелы на один
        import re
        text = re.sub(r'\s+', ' ', text)
        
        # Заменяем проблемные символы
        replacements = {
            '\n': ' ',      # Перевод строки
            '\r': ' ',      # Возврат каретки
            '\t': ' ',      # Табуляция
            '"': '"',       # Двойные кавычки (оставляем как есть)
            "'": "'",       # Одинарные кавычки (оставляем как есть)
            ';': ',',       # Точка с запятой
            '–': '-',       # Длинное тире
            '—': '-',       # Короткое тире
            '…': '...',     # Многоточие
        }
        
        for old_char, new_char in replacements.items():
            text = text.replace(old_char, new_char)
        
        return text
    
    def _determine_group_from_content(self, review_text: str, name: str) -> str:
        """
        Определяет группу объекта на основе содержимого
        
        Args:
            review_text: Текст отзыва
            name: Название объекта
            
        Returns:
            Определенная группа
        """
        # Ключевые слова для определения групп
        group_keywords = {
            'school': ['школа', 'учитель', 'ученик', 'класс', 'урок', 'директор', 'завуч'],
            'hospital': ['больница', 'врач', 'пациент', 'лечение', 'медицинский', 'клиника', 'поликлиника'],
            'pharmacy': ['аптека', 'лекарство', 'фармацевт', 'медикамент', 'препарат'],
            'kindergarden': ['детский сад', 'воспитатель', 'группа', 'игра', 'развитие'],
            'polyclinic': ['поликлиника', 'врач', 'прием', 'консультация', 'диспансер'],
            'university': ['университет', 'студент', 'преподаватель', 'лекция', 'сессия', 'факультет'],
            'shopmall': ['торговый центр', 'магазин', 'молл', 'гипермаркет', 'ТЦ', 'шопинг'],
            'resident_complexes': ['жилой комплекс', 'дом', 'квартира', 'жилье', 'недвижимость']
        }
        
        # Объединяем текст для анализа
        text_to_analyze = f"{review_text} {name}".lower()
        
        # Ищем совпадения
        for group, keywords in group_keywords.items():
            for keyword in keywords:
                if keyword in text_to_analyze:
                    return group
        
        return 'unknown'
    
    def process_excel_file(self, file_path: str, sheet_name: str = None, 
                          filters: Dict = None) -> pd.DataFrame:
        """
        Обработка Excel файла с возможностью фильтрации
        
        Args:
            file_path: Путь к Excel файлу
            sheet_name: Имя листа (если None, используется первый)
            filters: Словарь с фильтрами {поле: значение}
            
        Returns:
            DataFrame с данными
        """
        try:
            logger.info(f"Начинаем обработку Excel файла: {file_path}")
            
            # Получаем информацию о файле
            file_info = self.get_excel_info(file_path)
            logger.info(f"Информация о файле: {file_info}")
            
            if 'error' in file_info:
                logger.error(f"Ошибка чтения Excel файла: {file_info['error']}")
                return pd.DataFrame()
            
            # Определяем лист для чтения
            if sheet_name is None:
                if file_info['sheets']:
                    sheet_name = list(file_info['sheets'].keys())[0]
                    logger.info(f"Автоматически выбран лист: {sheet_name}")
                else:
                    logger.error("Не найден подходящий лист в Excel файле")
                    return pd.DataFrame()
            
            if sheet_name is None:
                logger.error("Не найден подходящий лист в Excel файле")
                return pd.DataFrame()
            
            # Читаем данные из указанного листа
            logger.info(f"Чтение листа '{sheet_name}' из файла {file_path}")
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            logger.info(f"Исходные колонки: {list(df.columns)}")
            logger.info(f"Загружено {len(df)} строк из листа '{sheet_name}'")
            
            if df.empty:
                logger.warning(f"Лист '{sheet_name}' пустой")
                return df
            
            # Применяем фильтры если указаны
            if filters:
                df = self._apply_filters(df, filters)
                logger.info(f"После применения фильтров осталось {len(df)} строк")
            
            # Применяем маппинг полей
            df_renamed = df.copy()
            for old_name, new_name in self.field_mapping.items():
                if old_name in df_renamed.columns:
                    df_renamed = df_renamed.rename(columns={old_name: new_name})
                    logger.info(f"Переименовано поле: {old_name} -> {new_name}")
            
            logger.info(f"Колонки после маппинга: {list(df_renamed.columns)}")
            
            # Фильтруем только поддерживаемые поля
            available_fields = [col for col in df_renamed.columns if col in self.supported_fields]
            logger.info(f"Поддерживаемые поля: {available_fields}")
            
            if not available_fields:
                logger.error("Не найдено ни одного поддерживаемого поля")
                logger.error(f"Доступные поля: {list(df_renamed.columns)}")
                logger.error(f"Поддерживаемые поля: {self.supported_fields}")
                return pd.DataFrame()
            
            # Оставляем только поддерживаемые поля
            df = df_renamed[available_fields]
            
            # Очищаем текстовые поля
            text_fields = ['review_text', 'answer_text', 'name', 'address', 'user_name']
            for field in text_fields:
                if field in df.columns:
                    df[field] = df[field].apply(self.clean_text_field)
            
            # Проверяем обязательные поля для обработки
            missing_required = [field for field in self.required_fields_processing 
                              if field not in df.columns]
            
            if missing_required:
                logger.error(f"Отсутствуют обязательные поля: {missing_required}")
                logger.error(f"Доступные поля: {list(df.columns)}")
                return pd.DataFrame()
            
            # Добавляем недостающие поля с пустыми значениями
            for field in self.supported_fields:
                if field not in df.columns:
                    if field in ['latitude', 'longitude']:
                        df[field] = 0.0
                    else:
                        df[field] = ""
            
            # Приводим к стандартному порядку полей
            df = df.reindex(columns=self.supported_fields)
            
            # Определяем determined_group, если он пустой
            for idx, row in df.iterrows():
                if not row.get('determined_group'):
                    group = self._determine_group_from_content(row.get('review_text', ''), row.get('name', ''))
                    df.at[idx, 'determined_group'] = group
            
            logger.info(f"Обработан Excel файл: {len(df)} строк")
            logger.info(f"Финальные колонки: {list(df.columns)}")
            return df
            
        except Exception as e:
            logger.error(f"Ошибка обработки Excel файла {file_path}: {str(e)}")
            import traceback
            logger.error(f"Полный traceback: {traceback.format_exc()}")
            return pd.DataFrame()
    
    def _apply_filters(self, df: pd.DataFrame, filters: Dict) -> pd.DataFrame:
        """
        Применение фильтров к DataFrame
        
        Args:
            df: DataFrame для фильтрации
            filters: Словарь с фильтрами {поле: значение}
            
        Returns:
            Отфильтрованный DataFrame
        """
        try:
            filtered_df = df.copy()
            
            for field, value in filters.items():
                if field in filtered_df.columns:
                    if isinstance(value, list):
                        # Фильтр по списку значений
                        filtered_df = filtered_df[filtered_df[field].isin(value)]
                    elif isinstance(value, dict):
                        # Сложный фильтр
                        if 'min' in value and 'max' in value:
                            # Диапазон значений
                            filtered_df = filtered_df[
                                (filtered_df[field] >= value['min']) & 
                                (filtered_df[field] <= value['max'])
                            ]
                        elif 'contains' in value:
                            # Содержит текст
                            filtered_df = filtered_df[
                                filtered_df[field].astype(str).str.contains(value['contains'], na=False)
                            ]
                    else:
                        # Точное совпадение
                        filtered_df = filtered_df[filtered_df[field] == value]
                    
                    logger.info(f"Применен фильтр {field}={value}: осталось {len(filtered_df)} строк")
            
            return filtered_df
            
        except Exception as e:
            logger.error(f"Ошибка применения фильтров: {str(e)}")
            return df
    
    def get_available_filters(self, file_path: str, sheet_name: str = None) -> Dict:
        """
        Получение доступных фильтров для Excel файла
        
        Args:
            file_path: Путь к Excel файлу
            sheet_name: Имя листа
            
        Returns:
            Словарь с доступными фильтрами
        """
        try:
            # Читаем данные для анализа
            if sheet_name is None:
                file_info = self.get_excel_info(file_path)
                sheet_name = list(file_info['sheets'].keys())[0] if file_info['sheets'] else None
            
            if sheet_name is None:
                return {}
            
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=1000)  # Ограничиваем для анализа
            
            filters = {}
            
            for column in df.columns:
                if column in self.supported_fields:
                    # Анализируем тип данных
                    if df[column].dtype in ['object', 'string']:
                        # Текстовые данные - предлагаем уникальные значения
                        unique_values = df[column].dropna().unique()
                        if len(unique_values) <= 20:  # Не более 20 значений
                            filters[column] = {
                                'type': 'select',
                                'values': unique_values.tolist(),
                                'sample_values': unique_values[:5].tolist()
                            }
                        else:
                            filters[column] = {
                                'type': 'text',
                                'description': f'Поиск по тексту в поле {column}'
                            }
                    elif df[column].dtype in ['int64', 'float64']:
                        # Числовые данные - предлагаем диапазон
                        min_val = df[column].min()
                        max_val = df[column].max()
                        filters[column] = {
                            'type': 'range',
                            'min': min_val,
                            'max': max_val,
                            'description': f'Диапазон значений для {column}'
                        }
            
            return filters
            
        except Exception as e:
            logger.error(f"Ошибка получения фильтров: {str(e)}")
            return {}
    
    def validate_dataframe(self, df: pd.DataFrame) -> Dict:
        """
        Валидация DataFrame
        
        Args:
            df: DataFrame для валидации
            
        Returns:
            Словарь с результатами валидации
        """
        if df.empty:
            return {
                'valid': False,
                'error': 'DataFrame пустой',
                'valid_records': 0,
                'invalid_records': 0,
                'missing_fields': []
            }
        
        # Проверяем обязательные поля для обработки
        missing_processing = [field for field in self.required_fields_processing 
                            if field not in df.columns]
        
        # Проверяем обязательные поля для архива
        missing_archive = [field for field in self.required_fields_archive 
                          if field not in df.columns]
        
        # Проверяем наличие данных в обязательных полях
        valid_mask = df[self.required_fields_processing].notna().all(axis=1)
        valid_records = df[valid_mask]
        invalid_records = df[~valid_mask]
        
        # Проверяем адреса для архива
        archive_mask = valid_records['address'].notna() & (valid_records['address'] != '')
        valid_for_archive = valid_records[archive_mask]
        addressless = valid_records[~archive_mask]
        
        result = {
            'valid': len(missing_processing) == 0 and len(valid_records) > 0,
            'valid_records': len(valid_records),
            'invalid_records': len(invalid_records),
            'valid_for_archive': len(valid_for_archive),
            'addressless_records': len(addressless),
            'missing_processing_fields': missing_processing,
            'missing_archive_fields': missing_archive,
            'total_records': len(df)
        }
        
        return result
    
    def convert_to_archive_format(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Приведение DataFrame к формату архива
        
        Args:
            df: Исходный DataFrame
            
        Returns:
            DataFrame в формате архива
        """
        # Создаем копию с нужными полями
        archive_df = pd.DataFrame()
        
        for field in self.archive_field_order:
            if field in df.columns:
                archive_df[field] = df[field]
            else:
                # Добавляем пустое поле
                if field in ['latitude', 'longitude']:
                    archive_df[field] = 0.0
                elif field in ['sentiment_score', 'positive_words_count', 'negative_words_count']:
                    archive_df[field] = 0.0
                else:
                    archive_df[field] = ""
        
        return archive_df 